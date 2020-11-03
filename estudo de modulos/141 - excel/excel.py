"""
https://openpyxl.readthedocs.io/en/stable/
pip install openpyxl
pipenv install openpyxl
"""
import openpyxl
from random import uniform
"""
pedidos = openpyxl.load_workbook('pedidos.xlsx')
nome_planilha = pedidos.sheetnames  # pra pegar quantas planilas tem no excel
planilha1 = pedidos['Página1']

# print(planilha1['b4'].value)
# print(10 * "=====")
# for linha in planilha1['a1:c2']:
#     for coluna in linha:
#         print(coluna.value)

print(10 * "=====")

# for linha in planilha1:
#     for coluna in linha:
#         print(coluna.value)


# for linha in planilha1:
#     if linha[0].value is not None:
#         print(linha[0].value, end=" ")
#
#     if linha[1].value is not None:
#         print(linha[1].value, end=" ")
#
#     if linha[2].value is not None:
#         print(linha[2].value)

#alterando os dados na planilha
#planilha1['b3'].value = 2200

for linha in range(5,16):
    num_pedido = linha - 1
    planilha1.cell(linha, 1).value = num_pedido
    planilha1.cell(linha, 2).value = 1200 + linha
    preco = round(uniform(10, 100), 2)
    planilha1.cell(linha, 3).value = preco

    print(linha)

pedidos.save('nova_planilha.xlsx')
"""
print('criando uma nova planilha')
planilha = openpyxl.Workbook()
planilha.create_sheet('Planilha1', 0)#nome da planilha + indice
planilha.create_sheet('Planilha2', 1)

planilha1 = planilha['Planilha1']
planilha2 = planilha['Planilha2']

for linha in range(1,11):
    num_pedido = linha - 1
    planilha1.cell(linha, 1).value = num_pedido
    planilha1.cell(linha, 2).value = 1200 + linha
    preco = round(uniform(10, 100), 2)
    planilha1.cell(linha, 3).value = preco

for linha in range(1,11):
    planilha2.cell(linha, 1).value = f'Luiz {linha} {round(uniform(10, 100))}'
    planilha2.cell(linha, 2).value = f'Otavio {linha} {round(uniform(10, 100))}'
    planilha2.cell(linha, 3).value = f'Mariazinha {linha} {round(uniform(10, 100))}'

planilha.save('nova_planilha2.xlsx')